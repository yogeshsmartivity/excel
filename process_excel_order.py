import os
import re
import sys
import argparse
import warnings
warnings.filterwarnings("ignore", category=FutureWarning)
import pandas as pd
import pypdf
import win32com.client

GITHUB_RAW_BASE = "https://raw.githubusercontent.com/yogeshsmartivity/excel/main/"
CURRENT_VERSION = "1.1.3"

_ver_txt = os.path.join(os.path.dirname(os.path.abspath(__file__)), "version.txt")
if os.path.exists(_ver_txt):
    try:
        with open(_ver_txt, "r", encoding="utf-8") as _vf:
            for _line in _vf:
                if _line.startswith("VERSION="):
                    CURRENT_VERSION = _line.split("=")[1].strip()
    except Exception:
        pass

def check_for_updates(workbook_path=None, force_download=False):
    """
    Checks GitHub Raw URL for online updates and downloads missing or updated files:
    1. process_excel_order.py
    2. github_api_push.py
    3. master_price_list.xlsx
    4. master_discount_list.xlsx
    5. version.txt
    """
    try:
        import urllib.request
        wb_dir = os.path.dirname(os.path.abspath(workbook_path)) if workbook_path else os.path.dirname(os.path.abspath(__file__))
        version_file = os.path.join(wb_dir, "version.txt")
        
        current_ver = CURRENT_VERSION
        if os.path.exists(version_file):
            try:
                with open(version_file, "r", encoding="utf-8") as vf:
                    for line in vf:
                        if line.startswith("VERSION="):
                            current_ver = line.strip().split("=")[1]
            except Exception:
                pass
                
        # Check for missing critical files locally
        missing_files = []
        for check_fn in ["github_api_push.py", "master_price_list.xlsx", "master_discount_list.xlsx"]:
            if not os.path.exists(os.path.join(wb_dir, check_fn)):
                missing_files.append(check_fn)
                
        print(f"System Version: v{current_ver} (GitHub Full Auto-Sync Enabled)")
        
        online_ver = current_ver
        try:
            online_ver_url = GITHUB_RAW_BASE + "version.txt"
            req = urllib.request.Request(online_ver_url, headers={'User-Agent': 'Mozilla/5.0'})
            with urllib.request.urlopen(req, timeout=5) as resp:
                online_text = resp.read().decode('utf-8')
                for l in online_text.split('\n'):
                    if l.startswith("VERSION="):
                        online_ver = l.strip().split("=")[1]
        except Exception as net_err:
            print(f"Could not reach GitHub for version check: {net_err}")

        should_update = force_download or (online_ver != current_ver) or (len(missing_files) > 0)

        if should_update:
            print(f"Downloading latest files from GitHub (v{online_ver})...")
            files_to_dl = [
                "process_excel_order.py",
                "github_api_push.py",
                "master_price_list.xlsx",
                "master_discount_list.xlsx",
                "version.txt"
            ]
            for fn in files_to_dl:
                try:
                    dl_url = GITHUB_RAW_BASE + fn
                    dl_path = os.path.join(wb_dir, fn)
                    urllib.request.urlretrieve(dl_url, dl_path)
                    print(f"  [DOWNLOADED] {fn}")
                except Exception as dl_err:
                    print(f"  [DL ERROR] {fn}: {dl_err}")
                    
            # Immediately sync master price list & discount sheets into active workbook
            if workbook_path and os.path.exists(workbook_path):
                try:
                    excel_app = None
                    try:
                        excel_app = win32com.client.GetActiveObject("Excel.Application")
                    except Exception:
                        excel_app = win32com.client.Dispatch("Excel.Application")
                        
                    if excel_app:
                        wb_active = None
                        try:
                            wb_active = excel_app.ActiveWorkbook
                        except Exception:
                            pass
                            
                        if not wb_active:
                            wb_active = excel_app.Workbooks.Open(workbook_path)
                            
                        if wb_active:
                            sync_master_price_list(wb_active, wb_dir)
                            wb_active.Save()
                except Exception as sync_active_err:
                    print(f"Active workbook sync note: {sync_active_err}")

            notice_path = os.path.join(wb_dir, "update_notice.txt")
            msg = f"🚀 LATEST PRICE LIST & DISCOUNTS INSTALLED FROM GITHUB!\n\nSystem Version: v{online_ver}\nAll master price lists, discount tables, and features have been updated in your active Excel workbook!\n\nClick OK to continue."
            try:
                with open(notice_path, "w", encoding="utf-8") as nf:
                    nf.write("STATUS=UPDATED\n")
                    nf.write(f"VERSION={online_ver}\n")
                    nf.write(f"MSG={msg}")
            except Exception:
                pass
                
            return True, online_ver
        else:
            return False, current_ver
            
    except Exception as e:
        print(f"Auto-update check skipped: {e}")
        return False, CURRENT_VERSION

def sync_master_price_list(wb, wb_dir):
    """
    If master_price_list.xlsx or master_discount_list.xlsx exists, syncs updated rows into Price list and discount sheets.
    """
    master_path = os.path.join(wb_dir, "master_price_list.xlsx")
    if os.path.exists(master_path):
        try:
            print("Syncing Master Price List into active workbook...")
            import openpyxl
            wb_master = openpyxl.load_workbook(master_path, data_only=True)
            if "Price list" in wb_master.sheetnames:
                sh_m = wb_master["Price list"]
                sh_price = wb.Sheets("Price list")
                for r in range(1, sh_m.max_row + 1):
                    for c in range(1, max(7, sh_m.max_column + 1)):
                        val = sh_m.cell(r, c).value
                        if val is not None:
                            try:
                                sh_price.Cells(r, c).Value = str(val) if not isinstance(val, (int, float)) else val
                            except Exception:
                                pass
                                
                # Clear trailing rows beyond master max_row
                try:
                    last_price_r = max(sh_price.Cells(sh_price.Rows.Count, "A").End(-4162).Row, sh_price.Cells(sh_price.Rows.Count, "B").End(-4162).Row)
                    if last_price_r > sh_m.max_row:
                        sh_price.Range(f"A{sh_m.max_row + 1}:K{last_price_r + 10}").ClearContents()
                        print(f"Cleared trailing deleted rows {sh_m.max_row + 1} to {last_price_r}.")
                except Exception as clr_err:
                    print(f"Note clearing trailing rows: {clr_err}")
                    
                print(f"Price list synced successfully ({sh_m.max_row} rows)!")
        except Exception as sync_err:
            print(f"Warning syncing master price list: {sync_err}")

    # Sync Master Discount List
    disc_master_path = os.path.join(wb_dir, "master_discount_list.xlsx")
    target_disc_path = disc_master_path if os.path.exists(disc_master_path) else master_path
    if os.path.exists(target_disc_path):
        try:
            import openpyxl
            wb_disc = openpyxl.load_workbook(target_disc_path, data_only=True)
            if "discount" in wb_disc.sheetnames:
                print("Syncing Master Discount List into active workbook...")
                sh_dm = wb_disc["discount"]
                sh_disc_wb = wb.Sheets("discount")
                for r in range(1, sh_dm.max_row + 1):
                    for c in range(1, max(7, sh_dm.max_column + 1)):
                        val = sh_dm.cell(r, c).value
                        if val is not None:
                            try:
                                sh_disc_wb.Cells(r, c).Value = str(val) if not isinstance(val, (int, float)) else val
                            except Exception:
                                pass
                print(f"Discount list synced successfully ({sh_dm.max_row} rows)!")
        except Exception as disc_err:
            print(f"Warning syncing discount list: {disc_err}")

def parse_pdf_with_gemini(file_path, api_key):
    import google.generativeai as genai
    import json
    
    print("Initializing Gemini API for AI-powered OCR...")
    genai.configure(api_key=api_key)
    
    # Upload the PDF file to Gemini API
    print(f"Uploading '{os.path.basename(file_path)}' to Gemini...")
    uploaded_file = genai.upload_file(file_path)
    
    # Use gemini-1.5-flash which is fast, free-tier supported, and natively handles PDF files.
    model = genai.GenerativeModel("gemini-1.5-flash")
    
    is_firstcry = "firstcry" in os.path.basename(file_path).lower()
    
    if is_firstcry:
        prompt = """
        Analyze this purchase order PDF.
        Extract the following information:
        1. The name of the buyer/party (e.g. DIGITAL AGE RETAIL PRIVATE LIMITED, etc.). Return this as 'party_name'.
        2. The list of items ordered. For each item, extract:
           - 'sku': The Product ID (e.g. 15433823).
           - 'name': The name/description of the item.
           - 'qty': The main ordered quantity (integer).
           - 'scheme': The free or scheme quantity if mentioned (integer, default 0).
           - 'mrp': The printed unit rate or MRP (float).
           
        Return the result strictly as a JSON object with two keys:
        {
          "party_name": "...",
          "items": [
            {
              "sku": "...",
              "name": "...",
              "qty": 10,
              "scheme": 0,
              "mrp": 100.0
            },
            ...
          ]
        }
        Make sure to output ONLY the raw JSON string. Do not wrap it in markdown backticks or blockquotes.
        """
    else:
        prompt = """
        Analyze this purchase order PDF.
        Extract the following information:
        1. The name of the buyer/party (e.g. APEX ENTERPRISES, BLINK COMMERCE PRIVATE LIMITED, etc.). Return this as 'party_name'.
        2. The list of items ordered. For each item, extract:
           - 'sku': The EAN/UPC barcode (13-digit number if visible) or the SKU code.
           - 'name': The name/description of the item.
           - 'qty': The main ordered quantity (integer).
           - 'scheme': The free or scheme quantity if mentioned (integer, default 0).
           - 'mrp': The printed unit rate or MRP (float).
           
        Return the result strictly as a JSON object with two keys:
        {
          "party_name": "...",
          "items": [
            {
              "sku": "...",
              "name": "...",
              "qty": 10,
              "scheme": 0,
              "mrp": 100.0
            },
            ...
          ]
        }
        Make sure to output ONLY the raw JSON string. Do not wrap it in markdown backticks or blockquotes.
        """
    
    print("Extracting fields using Gemini AI...")
    response = model.generate_content([uploaded_file, prompt])
    
    # Clean up the file from Gemini
    try:
        uploaded_file.delete()
    except Exception as del_err:
        print(f"VBA API Clean-up Warning: {del_err}")
        
    text = response.text.strip()
    # Clean output backticks if any
    if text.startswith("```json"):
        text = text[7:]
    elif text.startswith("```"):
        text = text[3:]
    if text.endswith("```"):
        text = text[:-3]
    text = text.strip()
    
    data = json.loads(text)
    party_name = data.get("party_name", "UNKNOWN PARTY").strip()
    extracted_items = data.get("items", [])
    
    # Make sure all elements in extracted_items have correct types
    formatted_items = []
    for item in extracted_items:
        sku_val = str(item.get("sku", "")).strip()
        if sku_val.endswith(".0"):
            sku_val = sku_val[:-2]
            
        formatted_items.append({
            'sku': sku_val,
            'name': str(item.get("name", "")).strip(),
            'qty': int(item.get("qty", 0)),
            'scheme': int(item.get("scheme", 0)),
            'mrp': float(item.get("mrp", 0.0)),
            'base_cost': float(item.get("base_cost", 0.0))
        })
        
    return party_name, formatted_items

def parse_pdf_order(file_path):
    print(f"Parsing PDF Order: {os.path.basename(file_path)}...")
    reader = pypdf.PdfReader(file_path)
    lines = []
    for page in reader.pages:
        text = page.extract_text()
        if text:
            lines.extend(text.split('\n'))
            
    party_name = "UNKNOWN PARTY"
    for idx, l in enumerate(lines):
        if "BILLTO" in l.upper().replace(" ", ""):
            if idx + 1 < len(lines):
                party_name = lines[idx + 1].strip()
                break
                
    extracted_items = []
    i = 0
    seen_skus = set()
    
    while i < len(lines):
        line = lines[i].strip()
        
        # Match serial number or SMRT line directly
        sr_match = re.match(r'^(\d+)(?:\s+(.*))?$', line)
        smrt_in_line = "SMRT" in line.upper()
        
        if sr_match or smrt_in_line:
            sr_num = None
            first_name_part = ""
            if sr_match:
                try:
                    sr_num = int(sr_match.group(1))
                    if sr_num > 150:
                        i += 1
                        continue
                except ValueError:
                    pass
                first_name_part = sr_match.group(2) if sr_match.group(2) else ""
                
            if first_name_part and any(k in first_name_part.upper() for k in ["REMARKS", "PENDING INVOICE", "FOC SCHEME", "INVOICE NO", "SUBTOTAL", "GRAND TOTAL"]):
                i += 1
                continue
                
            # Find the SMRT line within next 6 lines
            smrt_idx = -1
            if smrt_in_line:
                smrt_idx = i
            else:
                for k in range(i, min(i + 8, len(lines))):
                    if "SMRT" in lines[k].upper():
                        smrt_idx = k
                        break
                        
            if smrt_idx != -1:
                # Reconstruct product name
                name_parts = []
                if first_name_part and not any(k in first_name_part.upper() for k in ["FOC SCHEME", "PENDING INVOICE"]):
                    name_parts.append(first_name_part)
                for k in range(i + 1, smrt_idx):
                    part = lines[k].strip()
                    if not "SMRT" in part.upper() and not part.isdigit() and "BILL TO" not in part and "ORDER INFO" not in part and "FOC SCHEME" not in part and "PENDING INVOICE" not in part:
                        name_parts.append(part)
                        
                name_raw = " ".join(name_parts)
                name_clean = re.sub(r'^(.*?-\s*₹\d+\s*|\d+\s*)+', '', name_raw).strip()
                
                # Gather SMRT line and surrounding numeric lines
                smrt_line = lines[smrt_idx].strip()
                tokens = smrt_line.split()
                
                # Expand tokens by inspecting up to 4 lines forward
                for k_next in range(smrt_idx + 1, min(smrt_idx + 5, len(lines))):
                    nxt_str = lines[k_next].strip()
                    if nxt_str.isdigit() or "₹" in nxt_str or nxt_str.startswith("-"):
                        tokens.extend(nxt_str.split())
                    elif len(tokens) <= 2 and ("₹" in nxt_str or any(c.isdigit() for c in nxt_str)):
                        tokens.extend(nxt_str.split())
                        
                # Extract SMRT SKU token
                sku = None
                for t in tokens:
                    if "SMRT" in t.upper():
                        sku = re.sub(r'[^A-Za-z0-9-]', '', t.upper())
                        break
                if not sku:
                    sku = tokens[0]
                    
                # Find MRP token index
                mrp_idx = -1
                for t_idx, token in enumerate(tokens):
                    if '₹' in token or token.startswith('₹'):
                        mrp_idx = t_idx
                        break
                if mrp_idx == -1:
                    for t_idx, token in enumerate(tokens):
                        if t_idx > 0 and token.replace(',', '').replace('.', '').isdigit():
                            mrp_idx = t_idx
                            break
                            
                if mrp_idx != -1:
                    mrp_str = tokens[mrp_idx].replace('₹', '').replace(',', '').strip()
                    try:
                        mrp = float(mrp_str)
                    except ValueError:
                        mrp = 0.0
                        
                    qty = 0
                    scheme = 0
                    if mrp_idx + 1 < len(tokens):
                        try:
                            qty = int(tokens[mrp_idx + 1])
                        except ValueError:
                            pass
                    if mrp_idx + 2 < len(tokens):
                        try:
                            scheme = int(tokens[mrp_idx + 2])
                        except ValueError:
                            pass
                            
                    # Avoid duplicate extraction of same line
                    item_key = (sku, mrp, qty, scheme, smrt_idx)
                    if item_key not in seen_skus:
                        seen_skus.add(item_key)
                        extracted_items.append({
                            'sku': sku,
                            'name': name_clean,
                            'mrp': mrp,
                            'qty': qty,
                            'scheme': scheme
                        })
                        i = smrt_idx + 1
                        continue
        i += 1
    return party_name, extracted_items

def parse_blinkit_po(file_path):
    print(f"Parsing Blinkit PDF Order: {os.path.basename(file_path)}...")
    reader = pypdf.PdfReader(file_path)
    lines = []
    for page in reader.pages:
        text = page.extract_text()
        if text:
            for l in text.split('\n'):
                if l.strip():
                    lines.append(l.strip())
                    
    item_indices = [idx for idx, line in enumerate(lines) if line == "890801"]
    extracted_items = []
    
    for item_idx in item_indices:
        if item_idx + 2 >= len(lines):
            continue
        upc = "890801" + lines[item_idx+1] + lines[item_idx+2]
        
        # Extract description
        desc_parts = []
        curr_idx = item_idx + 3
        while curr_idx < len(lines):
            line = lines[curr_idx]
            if re.match(r'^\d+(\.\d+)?$', line):
                break
            desc_parts.append(line)
            curr_idx += 1
        description = " ".join(desc_parts)
        
        # Extract numbers
        numbers = []
        while curr_idx < len(lines):
            line = lines[curr_idx]
            if re.match(r'^\d+(\.\d+)?$', line) or line == ".":
                if line != ".":
                    numbers.append(line)
                curr_idx += 1
            else:
                break
                
        # Reconstruct split floats
        merged_nums = []
        n_idx = 0
        while n_idx < len(numbers):
            num = numbers[n_idx]
            if n_idx + 1 < len(numbers) and '.' in num:
                next_num = numbers[n_idx + 1]
                if len(next_num) == 1 and next_num.isdigit():
                    merged_nums.append(num + next_num)
                    n_idx += 2
                    continue
            merged_nums.append(num)
            n_idx += 1
            
        # Find Qty and MRP
        margin_idx = -1
        for j, num in enumerate(merged_nums):
            try:
                val = float(num)
                if abs(val - 45.0) < 0.1:
                    margin_idx = j
                    break
            except ValueError:
                pass
                
        if margin_idx != -1:
            try:
                qty = int(float(merged_nums[margin_idx - 2]))
                mrp = float(merged_nums[margin_idx - 1])
                extracted_items.append({
                    'name': description,
                    'sku': upc, # Temp store UPC in sku
                    'mrp': mrp,
                    'qty': qty,
                    'scheme': 0
                })
            except Exception as ex:
                print(f"Warning extracting fields: {ex}")
                
    party_name = "BLINK COMMERCE PRIVATE LIMITED"
    return party_name, extracted_items

def parse_firstcry_po(file_path):
    print(f"Parsing Firstcry PDF Order: {os.path.basename(file_path)}...")
    reader = pypdf.PdfReader(file_path)
    lines = []
    for page in reader.pages:
        text = page.extract_text()
        if text:
            for l in text.split('\n'):
                if l.strip():
                    lines.append(l.strip())
                    
    extracted_items = []
    i = 0
    expected_sr = 1
    
    while i < len(lines):
        line = lines[i].strip()
        if line.isdigit() and int(line) == expected_sr:
            style_idx = -1
            for j in range(i + 1, min(i + 25, len(lines))):
                if re.match(r'^SMRT\d+', lines[j]):
                    style_idx = j
                    break
            
            if style_idx != -1:
                style_code = lines[style_idx].strip()
                hsn = lines[style_idx + 1].strip() if style_idx + 1 < len(lines) else ""
                qty_str = lines[style_idx + 2].strip() if style_idx + 2 < len(lines) else "0"
                mrp_str = lines[style_idx + 3].strip() if style_idx + 3 < len(lines) else "0.0"
                
                product_id = ""
                for j in range(i + 1, style_idx):
                    val = lines[j].strip()
                    if val.isdigit() and 7 <= len(val) <= 9:
                        product_id = val
                        break
                
                desc_parts = []
                prod_id_idx = -1
                for j in range(i + 1, style_idx):
                    val = lines[j].strip()
                    if val == product_id:
                        prod_id_idx = j
                        break
                
                if prod_id_idx != -1:
                    for j in range(prod_id_idx + 1, style_idx):
                        desc_parts.append(lines[j].strip())
                description = " ".join(desc_parts).replace('\t', ' ').strip()
                
                try:
                    qty = int(qty_str)
                    mrp_clean = mrp_str.replace('₹', '').replace(',', '').strip()
                    mrp = float(mrp_clean)
                    
                    base_cost_str = lines[style_idx + 6].strip() if style_idx + 6 < len(lines) else "0.0"
                    base_cost_clean = base_cost_str.replace('₹', '').replace(',', '').strip()
                    base_cost = float(base_cost_clean)
                    
                    extracted_items.append({
                        'name': description,
                        'sku': product_id,
                        'mrp': mrp,
                        'qty': qty,
                        'scheme': 0,
                        'base_cost': base_cost
                    })
                    expected_sr += 1
                    i = style_idx + 4
                    continue
                except Exception as ex:
                    print(f"Error parsing item {expected_sr} near line {i}: {ex}")
        i += 1
        
    party_name = "Digital Age Retail Pvt. Ltd."
    return party_name, extracted_items

def parse_excel_order(file_path):
    print(f"Parsing Excel Order: {os.path.basename(file_path)}...")
    df = pd.read_excel(file_path)
    
    col_map = {col: str(col).strip().upper() for col in df.columns}
    df = df.rename(columns=col_map)
    
    sku_col = None
    qty_col = None
    scheme_col = None
    name_col = None
    mrp_col = None
    
    for original, clean in col_map.items():
        if clean in ['SKU', 'SKU CODE', 'ITEM CODE', 'ITEMCODE', 'BN ITEM NO', 'BNITEMNO', 'PRODUCT CODE']:
            sku_col = clean
        elif clean in ['QTY', 'QUANTITY', 'QTY ORDERED', 'ORDER QTY', 'QUANTITY ORDERED']:
            qty_col = clean
        elif clean in ['SCHEME', 'SCHEME QTY', 'FREE', 'FREE QTY', 'SCHEME QUANTITY']:
            scheme_col = clean
        elif clean in ['NAME', 'ITEM NAME', 'PRODUCT NAME', 'DESCRIPTION', 'PRODUCT DESCRIPTION']:
            name_col = clean
        elif clean in ['MRP', 'RATE', 'PRICE', 'UNIT PRICE', 'BASIC RATE', 'ORDER MRP']:
            mrp_col = clean
            
    if not sku_col and len(df.columns) > 0:
        sku_col = df.columns[0]
    if not qty_col and len(df.columns) > 1:
        qty_col = df.columns[1]
    if not scheme_col and len(df.columns) > 2:
        scheme_col = df.columns[2]
        
    extracted_items = []
    party_name = "UNKNOWN PARTY"
    
    for col in df.columns:
        for idx in range(min(5, len(df))):
            val = str(df.loc[idx, col]).upper()
            if "BILL TO" in val or "PARTY" in val or "CUSTOMER" in val:
                if idx + 1 < len(df):
                    party_name = str(df.loc[idx + 1, col]).strip()
                    break
                    
    if party_name == "UNKNOWN PARTY":
        filename = os.path.basename(file_path)
        party_name = os.path.splitext(filename)[0].replace("Order", "").replace("order", "").strip()
        
    for idx, row in df.iterrows():
        sku = str(row[sku_col]).strip() if sku_col and pd.notna(row[sku_col]) else ""
        if not sku or sku.upper() == 'NAN' or sku == "":
            continue
            
        qty = int(row[qty_col]) if qty_col and pd.notna(row[qty_col]) else 0
        scheme = int(row[scheme_col]) if scheme_col and pd.notna(row[scheme_col]) else 0
        name = str(row[name_col]).strip() if name_col and pd.notna(row[name_col]) else ""
        mrp = float(row[mrp_col]) if mrp_col and pd.notna(row[mrp_col]) else 0.0
        
        extracted_items.append({
            'sku': sku,
            'qty': qty,
            'scheme': scheme,
            'name': name,
            'mrp': mrp
        })
    return party_name, extracted_items

def get_active_workbook(workbook_path):
    excel = None
    try:
        excel = win32com.client.GetObject(Class="Excel.Application")
    except Exception:
        print("Excel is not running. Launching Excel Application...")
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = True
        
    try:
        excel.DisplayAlerts = False
    except Exception:
        pass
        
    wb = None
    target_name = os.path.basename(workbook_path).lower()
    for w in excel.Workbooks:
        if w.Name.lower() == target_name or w.FullName.lower() == workbook_path.lower():
            wb = w
            break
            
    if not wb:
        if os.path.exists(workbook_path):
            print(f"Opening workbook: {workbook_path}")
            wb = excel.Workbooks.Open(os.path.abspath(workbook_path))
        else:
            raise FileNotFoundError(f"Workbook '{target_name}' not found at '{workbook_path}'.")
            
    return excel, wb

def load_ean_mappings(wb):
    mappings = {}
    try:
        sh_price = wb.Sheets("Price list")
        last_row = sh_price.Cells(sh_price.Rows.Count, "B").End(-4162).Row
        for r in range(2, last_row + 1):
            sku = str(sh_price.Cells(r, 2).Value).strip()
            ean = str(sh_price.Cells(r, 7).Value).strip() # Column G (7) is EAN
            if ean and ean != "None" and sku and sku != "None":
                # Strip decimals if Excel imported it as float
                if ean.endswith(".0"):
                    ean = ean[:-2]
                mappings[ean] = sku
    except Exception as e:
        print(f"Warning: Could not load EAN mappings from Price list: {e}")
    return mappings

def load_firstcry_mappings(wb, workbook_path=None):
    mappings = {}
    
    # 1. Try to load from active win32com object
    try:
        sh_price = wb.Sheets("Price list")
        last_row = sh_price.Cells(sh_price.Rows.Count, "B").End(-4162).Row
        for r in range(2, last_row + 1):
            sku = str(sh_price.Cells(r, 2).Value).strip()
            fc_val = sh_price.Cells(r, 8).Value
            if fc_val is not None:
                fc_id = str(fc_val).strip()
                if fc_id.endswith(".0"):
                    fc_id = fc_id[:-2]
                if fc_id and fc_id != "None" and fc_id != "#N/A" and fc_id != "#VALUE!" and sku and sku != "None":
                    mappings[fc_id] = sku
    except Exception as e:
        print(f"Warning: Could not load Firstcry mappings via win32com: {e}")
        
    # 2. Fall back to pandas reading from disk if no mappings loaded
    if not mappings and workbook_path and os.path.exists(workbook_path):
        print("VBA COM returned no mappings. Falling back to reading cached values from disk via pandas...")
        try:
            import pandas as pd
            df = pd.read_excel(workbook_path, sheet_name="Price list")
            sku_col = None
            fc_col = None
            for col in df.columns:
                col_name = str(col).strip().upper()
                if col_name == 'SKU':
                    sku_col = col
                elif col_name == 'FIRSTCRY':
                    fc_col = col
                    
            if sku_col and fc_col:
                for idx, row in df.iterrows():
                    sku = str(row[sku_col]).strip()
                    fc_val = row[fc_col]
                    if pd.notna(fc_val):
                        fc_id = str(fc_val).strip()
                        if fc_id.endswith(".0"):
                            fc_id = fc_id[:-2]
                        if fc_id and fc_id.lower() != 'nan' and fc_id != '#n/a' and fc_id != '#value!' and sku and sku.lower() != 'nan':
                            mappings[fc_id] = sku
        except Exception as pd_err:
            print(f"Warning: Could not load Firstcry mappings from disk file: {pd_err}")
            
    return mappings

def parse_swiggy_po(pdf_path):
    print(f"Parsing Swiggy PDF Order: {os.path.basename(pdf_path)}...")
    reader = pypdf.PdfReader(pdf_path)
    lines = []
    for page in reader.pages:
        txt = page.extract_text()
        if txt:
            lines.extend(txt.split('\n'))
            
    party_name = "CLOUDKART VENTURES PRIVATE LIMITED"
    for idx, l in enumerate(lines):
        if "BILLING ADDRESS" in l.upper():
            if idx + 1 < len(lines):
                party_name = lines[idx+1].strip()
                break
                
    extracted_items = []
    i = 0
    while i < len(lines):
        line = lines[i].strip()
        match_start = re.match(r'^(\d+)\s+(\d+)\s+(.*)', line)
        if match_start:
            swiggy_code = match_start.group(2)
            name_start = match_start.group(3)
            
            summary_idx = -1
            for k in range(i, min(i + 15, len(lines))):
                k_line = lines[k].strip()
                tokens = k_line.split()
                if len(tokens) >= 10 and tokens[0].isdigit() and len(tokens[0]) >= 6:
                    summary_idx = k
                    break
                    
            if summary_idx != -1:
                desc_parts = [name_start]
                for k in range(i + 1, summary_idx):
                    part = lines[k].strip()
                    if not part.startswith('Colour:') and not part.startswith('Size:') and not part.startswith('Brand:'):
                        desc_parts.append(part)
                raw_name = ' '.join(desc_parts)
                cleaned_name = re.sub(r'\s*\d+\.\d+\s+pack.*', '', raw_name, flags=re.IGNORECASE).strip()
                
                sum_tokens = lines[summary_idx].strip().split()
                hsn = sum_tokens[0]
                qty = int(sum_tokens[1])
                mrp = float(sum_tokens[2])
                base_cost = float(sum_tokens[3])
                
                extracted_items.append({
                    'name': cleaned_name,
                    'sku': swiggy_code,
                    'mrp': mrp,
                    'qty': qty,
                    'scheme': 0,
                    'base_cost': base_cost,
                    'hsn': hsn
                })
                i = summary_idx + 1
                continue
        i += 1
        
    return party_name, extracted_items

def detect_vendor(order_path):
    filename = os.path.basename(order_path).lower()
    if "firstcry" in filename:
        return "Firstcry Order"
    if "blinkit" in filename:
        return "Blinkit Order"
    if "swiggy" in filename or "cloudkart" in filename:
        return "Swiggy Order"
    if "amit" in filename:
        return "Amit Order"
        
    # Read PDF text to detect by content keywords
    ext = os.path.splitext(order_path)[1].lower()
    if ext == '.pdf':
        try:
            reader = pypdf.PdfReader(order_path)
            if len(reader.pages) > 0:
                text = reader.pages[0].extract_text()
                if text:
                    text_upper = text.upper()
                    if "DIGITAL AGE RETAIL" in text_upper or "FIRSTCRY" in text_upper:
                        return "Firstcry Order"
                    if "BLINK COMMERCE" in text_upper or "BLINKIT" in text_upper:
                        return "Blinkit Order"
                    if "CLOUDKART" in text_upper or "SWIGGY" in text_upper:
                        return "Swiggy Order"
                    if "APEX ENTERPRISES" in text_upper or "SMARTIVITY LABS" in text_upper:
                        return "Amit Order"
        except Exception as e:
            print(f"Warning during vendor auto-detection: {e}")
            
    # Default fallback
    return "Amit Order"

def run_import(order_path, workbook_path):
    print(f"Running Import to workbook: {workbook_path}...")
    
    excel, wb = get_active_workbook(workbook_path)
    
    # Auto-detect vendor sheet
    target_sheet_name = detect_vendor(order_path)
    print(f"Auto-detected target sheet for order: '{target_sheet_name}'")
    active_sheet_name = target_sheet_name
    sh_order = wb.Sheets(active_sheet_name)
    try:
        sh_order.Activate()
    except Exception:
        pass
    
    # 1. Parse Order File
    ext = os.path.splitext(order_path)[1].lower()
    is_blinkit = (active_sheet_name == "Blinkit Order" or "blinkit" in os.path.basename(order_path).lower())
    is_firstcry = (active_sheet_name == "Firstcry Order" or "firstcry" in os.path.basename(order_path).lower())
    is_swiggy = (active_sheet_name == "Swiggy Order" or "swiggy" in os.path.basename(order_path).lower() or "cloudkart" in os.path.basename(order_path).lower() or "cvpl" in os.path.basename(order_path).lower())
    
    if ext == '.pdf':
        # Inspect PDF text content to auto-detect vendor format
        pdf_content_text = ""
        try:
            reader_check = pypdf.PdfReader(order_path)
            for p_check in reader_check.pages[:2]:
                t_check = p_check.extract_text()
                if t_check:
                    pdf_content_text += t_check.upper() + " "
        except Exception:
            pass

        if "CLOUDKART" in pdf_content_text or "CHCPO" in pdf_content_text or "SWIGGY" in pdf_content_text:
            is_swiggy = True
        elif "BLINK COMMERCE" in pdf_content_text or "BLINKIT" in pdf_content_text:
            is_blinkit = True
        elif "FIRSTCRY" in pdf_content_text or "DIGITAL AGE RETAIL" in pdf_content_text:
            is_firstcry = True

        parsed_via_local = False
        items = []
        party_name = "UNKNOWN PARTY"
        
        # Try local parsing first
        try:
            if is_blinkit:
                party_name, items = parse_blinkit_po(order_path)
            elif is_firstcry:
                party_name, items = parse_firstcry_po(order_path)
            elif is_swiggy:
                party_name, items = parse_swiggy_po(order_path)
            else:
                party_name, items = parse_pdf_order(order_path)
            
            if items:
                parsed_via_local = True
                print("Successfully parsed PDF using local parser!")
        except Exception as local_err:
            print(f"Local parser failed: {local_err}")
            
        # Fallback to Gemini if local parsing failed
        if not parsed_via_local:
            print("Local parser failed. Checking if Gemini API can parse it...")
            wb_dir = os.path.dirname(os.path.abspath(workbook_path))
            key_file_path = os.path.join(wb_dir, "gemini_key.txt")
            api_key = None
            if os.path.exists(key_file_path):
                try:
                    with open(key_file_path, "r", encoding="utf-8") as kf:
                        api_key = kf.read().strip()
                except Exception as k_err:
                    print(f"Warning: Could not read gemini_key.txt: {k_err}")
                    
            if api_key and len(api_key) > 10 and not api_key.startswith("YOUR_GEMINI_API_KEY"):
                try:
                    party_name, items = parse_pdf_with_gemini(order_path, api_key)
                    if items:
                        print("Successfully parsed PDF using Gemini AI!")
                        # Write warning file to trigger Excel popup
                        warning_file_path = os.path.join(wb_dir, "format_changed_warning.txt")
                        try:
                            with open(warning_file_path, "w", encoding="utf-8") as wf:
                                wf.write("Local parser failed. Gemini AI was used to parse the file.")
                            print(f"Warning file created at: {warning_file_path}")
                        except Exception as wf_err:
                            print(f"Warning: Could not write warning file: {wf_err}")
                    else:
                        raise ValueError("Gemini returned zero items.")
                except Exception as gem_err:
                    print(f"Gemini API execution failed: {gem_err}")
                    raise RuntimeError("Failed to parse the PDF using both local parsers and Gemini AI.")
            else:
                raise RuntimeError("Local parser failed and no valid Gemini API key was found.")
    elif ext in ['.xls', '.xlsx']:
        party_name, items = parse_excel_order(order_path)
    else:
        print(f"Error: Unsupported file format '{ext}'")
        return
        
    print(f"Extracted Party Name: '{party_name}'")
    print(f"Extracted {len(items)} items.")
    
    if not items:
        print("Error: No items found in order file.")
        return
        
    # Load Price list references for Fuzzy SKU Match and Auto Price Correction
    import difflib
    wb_dir = os.path.dirname(os.path.abspath(workbook_path))
    sync_master_price_list(wb, wb_dir)
    
    print("Loading Price list references for smart matching and correction...")
    sh_price = wb.Sheets("Price list")
    last_price_row = max(sh_price.Cells(sh_price.Rows.Count, col).End(-4162).Row for col in ["B", "F", "G", "H", "I"])
    
    id_to_std_sku = {}
    base_to_std_sku = {}
    sku_prices = {}
    
    for r in range(2, last_price_row + 1):
        std_sku = str(sh_price.Cells(r, 2).Value).strip()
        if not std_sku or std_sku.lower() == "none":
            continue
            
        clean_sku = str(sh_price.Cells(r, 6).Value).strip().upper().replace(" ", "")
        ean = str(sh_price.Cells(r, 7).Value).strip().upper().replace(" ", "")
        fc_id = str(sh_price.Cells(r, 8).Value).strip().upper().replace(" ", "")
        swiggy_val = sh_price.Cells(r, 9).Value
        swiggy_id = str(swiggy_val).strip().upper().replace(" ", "") if swiggy_val is not None else ""
        
        if ean.endswith(".0"):
            ean = ean[:-2]
        if fc_id.endswith(".0"):
            fc_id = fc_id[:-2]
        if swiggy_id.endswith(".0"):
            swiggy_id = swiggy_id[:-2]
            
        price_val = sh_price.Cells(r, 3).Value
        try:
            sku_prices[std_sku] = float(price_val) if price_val is not None else 0.0
        except ValueError:
            sku_prices[std_sku] = 0.0
            
        id_to_std_sku[std_sku.upper().replace(" ", "")] = std_sku
        if clean_sku and clean_sku != "NONE":
            id_to_std_sku[clean_sku] = std_sku
            if clean_sku.startswith("SMRT") and len(clean_sku) >= 8:
                base_to_std_sku[clean_sku[:8]] = std_sku
        if ean and ean != "NONE":
            id_to_std_sku[ean] = std_sku
        if fc_id and fc_id != "NONE" and fc_id != "#N/A":
            id_to_std_sku[fc_id] = std_sku
        if swiggy_id and swiggy_id != "NONE" and swiggy_id != "#N/A":
            id_to_std_sku[swiggy_id] = std_sku
            
    # Resolve, Base Match and Correct Prices for all items
    for item in items:
        raw_sku = str(item['sku']).strip()
        if raw_sku.endswith(".0"):
            raw_sku = raw_sku[:-2]
            
        clean_key = raw_sku.upper().replace(" ", "")
        resolved_sku = None
        auto_matched = False
        
        if clean_key in id_to_std_sku:
            resolved_sku = id_to_std_sku[clean_key]
        elif clean_key.startswith("SMRT") and len(clean_key) >= 8:
            base_key = clean_key[:8]
            if base_key in base_to_std_sku:
                resolved_sku = base_to_std_sku[base_key]
                auto_matched = True
                print(f"Base SKU matched '{raw_sku}' to '{resolved_sku}'")
            else:
                resolved_sku = raw_sku
        else:
            resolved_sku = raw_sku
                
        item['sku'] = resolved_sku
        item['auto_matched_sku'] = auto_matched
        item['original_sku_po'] = raw_sku
        
        original_price = item['mrp']
        price_corrected = False
        
        if resolved_sku in sku_prices:
            list_price = sku_prices[resolved_sku]
            if list_price > 0.0 and abs(original_price - list_price) > 0.1:
                item['mrp'] = list_price
                price_corrected = True
                print(f"Price Auto-Corrected for {resolved_sku} (PO: {original_price} -> List: {list_price})")
                
        item['original_mrp_po'] = original_price
        item['price_corrected'] = price_corrected
            
    # 3. Clear existing table
    print(f"Clearing old data on {active_sheet_name} sheet...")
    sh_order.Range("C4").Value = ""
    sh_order.Range("C5").Value = ""
    
    last_row = sh_order.Cells(sh_order.Rows.Count, "A").End(-4162).Row # -4162 = xlUp
    if last_row >= 11:
        sh_order.Range(f"A11:D{last_row}").ClearContents()
        sh_order.Range(f"E11:M{last_row}").ClearContents()
        
    # 4. Populate table and write formulas dynamically
    print("Writing new items and formulas...")
    sh_order.Range("C4").Value = party_name
    sh_order.Range("C5").Value = os.path.basename(order_path)
    
    # Load Party Discounts for smart matching
    sh_disc = wb.Sheets("discount")
    last_disc_row = sh_disc.Cells(sh_disc.Rows.Count, "A").End(-4162).Row
    party_discounts = {}
    for dr in range(2, last_disc_row + 1):
        pname = str(sh_disc.Cells(dr, 1).Value or "").strip().upper()
        pclean = str(sh_disc.Cells(dr, 2).Value or "").strip().upper()
        d5 = float(sh_disc.Cells(dr, 3).Value or 0.0)
        d12 = float(sh_disc.Cells(dr, 4).Value or 0.0)
        d18 = float(sh_disc.Cells(dr, 5).Value or 0.0)
        d28 = float(sh_disc.Cells(dr, 6).Value or 0.0)
        disc_dict = {0.05: d5, 0.12: d12, 0.18: d18, 0.28: d28}
        if pname and pname != "NONE":
            party_discounts[pname] = disc_dict
        if pclean and pclean != "NONE":
            party_discounts[pclean] = disc_dict

    # Find matching party discount dict
    matched_party_disc = None
    clean_party_upper = party_name.upper().strip()
    for pk, pd_dict in party_discounts.items():
        if pk in clean_party_upper or clean_party_upper in pk:
            matched_party_disc = pd_dict
            break

    for idx, item in enumerate(items):
        r = 11 + idx
        sh_order.Cells(r, 1).Value = item['sku']
        sh_order.Cells(r, 2).Value = item['name']
        sh_order.Cells(r, 3).Value = item['qty']
        sh_order.Cells(r, 4).Value = item['scheme']
        sh_order.Cells(r, 5).Value = item['mrp']
        
        # Write Excel formulas dynamically
        match_term = f"LEFT(UPPER(SUBSTITUTE(A{r}, \" \", \"\")), 8) & \"*\""
        sh_order.Cells(r, 6).Value = f'=IF(ISBLANK(A{r}), "", IFERROR(INDEX(\'Price list\'!B:B, MATCH({match_term}, \'Price list\'!F:F, 0)), ""))'
        sh_order.Cells(r, 7).Value = f'=IF(ISBLANK(A{r}), "", IFERROR(INDEX(\'Price list\'!C:C, MATCH({match_term}, \'Price list\'!F:F, 0)), 0))'
        sh_order.Cells(r, 8).Value = f'=IF(ISBLANK(A{r}), "", IFERROR(INDEX(\'Price list\'!E:E, MATCH({match_term}, \'Price list\'!F:F, 0)), 0))'
        sh_order.Cells(r, 9).Value = f'=IF(ISBLANK(A{r}), "", IFERROR(INDEX(\'Price list\'!D:D, MATCH({match_term}, \'Price list\'!F:F, 0)), ""))'
        
        if active_sheet_name == "Firstcry Order":
            base_cost = item.get('base_cost', 0.0)
            sh_order.Cells(r, 10).Value = f'=IF(ISBLANK(A{r}), 0, IF(G{r}=0, 0, ROUND(1 - ({base_cost} / G{r}), 5)))'
            sh_order.Cells(r, 10).Value = f'=IF(ISBLANK(A{r}), "", IFERROR(IF(ROUND(H{r},2)=0.05, INDEX(discount!C:C, MATCH($C$4, discount!A:A, 0)), IF(ROUND(H{r},2)=0.18, INDEX(discount!E:E, MATCH($C$4, discount!A:A, 0)), IF(ROUND(H{r},2)=0.12, INDEX(discount!D:D, MATCH($C$4, discount!A:A, 0)), IF(ROUND(H{r},2)=0.28, INDEX(discount!F:F, MATCH($C$4, discount!A:A, 0)), INDEX(discount!C:C, MATCH($C$4, discount!A:A, 0)))))), 0))'
            
        sh_order.Cells(r, 11).Value = f'=IF(ISBLANK(A{r}), "", ROUND(G{r}*C{r}*J{r},2))'
        try:
            sh_order.Cells(r, 12).Value = f'=IF(ISBLANK(A{r}), "", (G{r}*C{r})-K{r})'
        except Exception:
            pass
        
        # Build status note based on fuzzy match & price correction
        status_note = ""
        if item.get('auto_matched_sku') and item.get('price_corrected'):
            status_note = f"✅ Auto-matched SKU & Price (PO: '{item['original_sku_po']}', \u20b9{item['original_mrp_po']:.2f})"
        elif item.get('auto_matched_sku'):
            status_note = f"✅ Auto-matched SKU (PO: '{item['original_sku_po']}')"
        elif item.get('price_corrected'):
            status_note = f"✅ Price Auto-Corrected (PO: \u20b9{item['original_mrp_po']:.2f})"
            
        if status_note:
            sh_order.Cells(r, 13).Value = status_note
        else:
            sh_order.Cells(r, 13).Value = f'=IF(ISBLANK(A{r}), "", IF(F{r}="", "⚠️ SKU not found in Price List", IF(ROUND(E{r},2)<>ROUND(G{r},2), "⚠️ Price Mismatch! (PO: " & TEXT(E{r},"#,##0.00") & ", List: " & TEXT(G{r},"#,##0.00") & ")", "✅ OK")))'
        
    # 5. Write GRAND TOTAL summary row at the bottom of the table
    if len(items) > 0:
        last_item_row = 10 + len(items)
        tot_row = last_item_row + 1
        
        sh_order.Range(f"A{tot_row}:B{tot_row}").Merge()
        sh_order.Range(f"A{tot_row}").Value = "GRAND TOTAL"
        sh_order.Range(f"A{tot_row}").HorizontalAlignment = -4108 # xlCenter
        
        sh_order.Cells(tot_row, 3).Value = f"=SUM(C11:C{last_item_row})"
        sh_order.Cells(tot_row, 4).Value = f"=SUM(D11:D{last_item_row})"
        sh_order.Cells(tot_row, 11).Value = f"=SUM(K11:K{last_item_row})"
        sh_order.Cells(tot_row, 12).Value = f"=SUM(L11:L{last_item_row})"
        
        tot_range = sh_order.Range(f"A{tot_row}:L{tot_row}")
        tot_range.Font.Name = "Segoe UI"
        tot_range.Font.Size = 11
        tot_range.Font.Bold = True
        tot_range.Interior.Color = 14803425 # Light Slate #CBD5E1
        tot_range.Borders.Weight = 3 # Thick border
        
    try:
        excel.Calculate()
        print("Excel formulas recalculated successfully.")
    except Exception as calc_err:
        print(f"Warning: Could not trigger Excel calculation: {calc_err}")
        
    # 6. Audit rows for Missing SKUs, Zero Prices, or Price Mismatches & Popup Alert
    issues = []
    if len(items) > 0:
        last_item_row = 10 + len(items)
        for r in range(11, last_item_row + 1):
            sku_val = str(sh_order.Cells(r, 1).Value or "").strip()
            name_val = str(sh_order.Cells(r, 2).Value or "").strip()
            plist_sku = str(sh_order.Cells(r, 6).Value or "").strip()
            
            try:
                base_price = float(sh_order.Cells(r, 7).Value or 0.0)
            except (ValueError, TypeError):
                base_price = 0.0
                
            try:
                mrp_val = float(sh_order.Cells(r, 5).Value or 0.0)
            except (ValueError, TypeError):
                mrp_val = 0.0
                
            # Check 1: Missing SKU in Price list
            if not plist_sku or plist_sku in ["", "#N/A", "#VALUE!"]:
                try:
                    sh_order.Range(f"A{r}:F{r}").Interior.Color = 9498110 # Light Yellow (#FEF08A)
                except Exception:
                    pass
                issues.append(f"• Row {r}: SKU '{sku_val}' ({name_val}) — Not found in Price List!")
            # Check 2: Zero Base Price
            elif base_price <= 0:
                try:
                    sh_order.Range(f"G{r}").Interior.Color = 9498110 # Light Yellow
                except Exception:
                    pass
                issues.append(f"• Row {r}: Base Price is \u20b90.00 for '{sku_val}'!")
            # Check 3: Price Mismatch
            elif abs(mrp_val - base_price) > 0.01 and mrp_val > 0:
                try:
                    sh_order.Range(f"E{r}").Interior.Color = 11181822 # Light Orange (#FED7AA)
                    sh_order.Range(f"G{r}").Interior.Color = 11181822
                except Exception:
                    pass
                issues.append(f"• Row {r}: Price Mismatch for '{sku_val}' (Order MRP: \u20b9{mrp_val:.2f}, Price List: \u20b9{base_price:.2f})")

    if issues:
        warning_msg = f"⚠️ ATTENTION: {len(issues)} ITEM(S) NEED REVIEW!\n\n"
        warning_msg += "\n".join(issues[:10])
        if len(issues) > 10:
            warning_msg += f"\n... and {len(issues) - 10} more items."
        warning_msg += "\n\nNote: Highlighted in Yellow/Orange cells on sheet for review."
        
        print("\n" + "="*50)
        print(warning_msg)
        print("="*50 + "\n")
        
        try:
            excel.MsgBox(warning_msg, 48, "Smartivity Order Import Warnings")
        except Exception:
            pass

    print("Import and verification completed successfully!")

def run_fill(workbook_path):
    print(f"Reading verified order rows from workbook: {workbook_path}...")
    excel, wb = get_active_workbook(workbook_path)
    
    active_sheet_name = None
    try:
        curr_name = excel.ActiveSheet.Name
        if curr_name in ["Amit Order", "Blinkit Order", "Firstcry Order", "Swiggy Order"] and wb.Sheets(curr_name).Range("C5").Value:
            active_sheet_name = curr_name
    except Exception:
        pass
        
    if not active_sheet_name:
        for sname in ["Amit Order", "Blinkit Order", "Firstcry Order", "Swiggy Order"]:
            if wb.Sheets(sname).Range("C5").Value:
                active_sheet_name = sname
                break
                
    if not active_sheet_name:
        active_sheet_name = "Amit Order"
        
    print(f"Reading from order sheet: '{active_sheet_name}'")
    sh_order = wb.Sheets(active_sheet_name)
    sh_temp = wb.Sheets("template")
    
    try:
        sh_order.Calculate()
        excel.Calculate()
        print("Excel formulas recalculated successfully.")
    except Exception as calc_err:
        print(f"Warning during recalculation: {calc_err}")
        
    party_name = sh_order.Range("C5").Value
    order_file = sh_order.Range("C6").Value
    if not party_name:
        print("Error: No order loaded in the sheet.")
        return
        
    # 1. Read Order Rows
    last_row = sh_order.Cells(sh_order.Rows.Count, "A").End(-4162).Row
    if last_row < 11:
        print("Error: No items found in the Order table.")
        return
        
    rows_to_fill = []
    print(f"Reading rows from row 11 to {last_row}...")
    
    for r in range(11, last_row + 1):
        sku_val = sh_order.Cells(r, 6).Value # Column F is Price List SKU
        sku = str(sku_val).strip() if sku_val is not None else ""
        if not sku or sku.upper() == 'NONE' or sku == "":
            sku_val = sh_order.Cells(r, 1).Value # Fallback to Column A (SMRT SKU)
            sku = str(sku_val).strip() if sku_val is not None else ""
            
        if not sku or sku.upper() == 'NONE' or sku == "":
            continue
            
        qty = int(sh_order.Cells(r, 3).Value) if sh_order.Cells(r, 3).Value else 0
        scheme = int(sh_order.Cells(r, 4).Value) if sh_order.Cells(r, 4).Value else 0
        
        # Read looked-up values
        rate = float(sh_order.Cells(r, 7).Value) if sh_order.Cells(r, 7).Value else 0.0
        tax = float(sh_order.Cells(r, 8).Value) if sh_order.Cells(r, 8).Value else 0.0
        discount_pct = float(sh_order.Cells(r, 10).Value) if sh_order.Cells(r, 10).Value else 0.0
        discount_amt = float(sh_order.Cells(r, 11).Value) if sh_order.Cells(r, 11).Value else 0.0
        
        rows_to_fill.append({
            'sku': sku,
            'qty': qty,
            'scheme': scheme,
            'rate': rate,
            'tax': tax,
            'discount_pct': discount_pct,
            'discount_amt': discount_amt
        })
        
    if active_sheet_name == "Firstcry Order":
        aggregated = {}
        for rinfo in rows_to_fill:
            key = (str(rinfo['sku']).strip(), round(rinfo['discount_pct'], 5))
            if key not in aggregated:
                aggregated[key] = {
                    'sku': rinfo['sku'],
                    'qty': 0,
                    'scheme': 0,
                    'rate': rinfo['rate'],
                    'tax': rinfo['tax'],
                    'discount_pct': rinfo['discount_pct'],
                    'discount_amt': 0.0
                }
            aggregated[key]['qty'] += rinfo['qty']
            aggregated[key]['scheme'] += rinfo['scheme']
            
        for agg_info in aggregated.values():
            agg_info['discount_amt'] = round(agg_info['rate'] * agg_info['qty'] * agg_info['discount_pct'], 2)
            
        rows_to_fill = list(aggregated.values())
        print(f"Aggregated Firstcry rows: {len(rows_to_fill)} unique rows.")
        
    if not rows_to_fill:
        print("Error: No valid rows to process.")
        return
        
    # 2. Clear Template Sheet
    print("Clearing template sheet...")
    last_temp_row = sh_temp.Cells(sh_temp.Rows.Count, "A").End(-4162).Row
    if last_temp_row >= 2:
        sh_temp.Range(f"A2:K{last_temp_row}").ClearContents()
        
    # 3. Build output rows (interleaving main and scheme rows)
    template_rows = []
    
    for item in rows_to_fill:
        # Add main item row
        if item['qty'] > 0:
            disc_pct = round(item['discount_pct'] * 100.0, 4) if item['discount_pct'] > 0 else 0.0
            
            template_rows.append([
                item['sku'],
                "", # VariantCodeNo
                item['qty'],
                item['rate'],
                disc_pct,
                "", # DiscountAmount (blank as requested)
                item['tax'] * 100.0 if item['tax'] > 0 else 0.0,
                "", # CustomerItemCode
                "", # CustomerItemName
                "", # ItemRemark
                ""  # InputField
            ])
            
        # Add scheme item row (placed immediately below main item row)
        if item['scheme'] > 0:
            rate_val = item['rate']
            scheme_qty = item['scheme']
            
            template_rows.append([
                item['sku'],
                "", # VariantCodeNo
                scheme_qty,
                rate_val, # Price is the base rate (not 0.0)
                100.0, # 100% discount
                "", # DiscountAmount (blank as requested)
                item['tax'] * 100.0 if item['tax'] > 0 else 0.0,
                "a", # CustomerItemCode
                "",
                "",
                ""
            ])
            
    # 4. Write to Template Sheet
    print(f"Writing {len(template_rows)} rows to template sheet...")
    for idx, row_val in enumerate(template_rows):
        r_temp = 2 + idx
        for col_idx, val in enumerate(row_val, 1):
            sh_temp.Cells(r_temp, col_idx).Value = val
            
    # 5. Auto-Export Template to separate xlsx workbook silently
    try:
        base_name = os.path.splitext(os.path.basename(str(order_file)))[0] if order_file else "Order"
        default_filename = f"{base_name}_mapped.xlsx"
        export_path = os.path.join(os.path.dirname(os.path.abspath(workbook_path)), default_filename)
        
        if os.path.exists(export_path):
            try:
                os.remove(export_path)
            except Exception:
                pass
                
        print(f"Exporting template silently to: {export_path}...")
        sh_temp.Copy()
        new_wb = excel.ActiveWorkbook
        new_wb.SaveAs(os.path.abspath(export_path), 51) # 51 = xlOpenXMLWorkbook (.xlsx)
        new_wb.Close(False)
        print("Template exported successfully!")
    except Exception as exp_err:
        print(f"Warning during auto-export: {exp_err}")
        
    try:
        wb.Save()
    except Exception:
        pass
        
    print("Template populated successfully!")

def push_team_masters(workbook_path):
    wb_dir = os.path.dirname(os.path.abspath(workbook_path)) if workbook_path else os.path.dirname(os.path.abspath(__file__))
    print("Exporting active workbook sheets to Master Excel files...")
    
    # Read live unsaved cells directly from active Excel COM window
    excel_data = {}
    try:
        excel = win32com.client.GetActiveObject("Excel.Application")
        wb_active = excel.ActiveWorkbook
        for sheet_name in ["Price list", "discount"]:
            if sheet_name in [s.Name for s in wb_active.Sheets]:
                sh = wb_active.Sheets(sheet_name)
                last_r = max(sh.Cells(sh.Rows.Count, "A").End(-4162).Row, sh.Cells(sh.Rows.Count, "B").End(-4162).Row)
                last_c = sh.Cells(1, sh.Columns.Count).End(-4159).Column
                if last_c < 6:
                    last_c = 6
                grid = []
                for r in range(1, last_r + 1):
                    row_v = [sh.Cells(r, c).Value for c in range(1, last_c + 1)]
                    grid.append(row_v)
                excel_data[sheet_name] = grid
    except Exception as com_err:
        print(f"Note: Excel COM live read note ({com_err}), falling back to openpyxl disk read...")
        
    import openpyxl
    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    
    # Export master_price_list.xlsx
    master_price_path = os.path.join(wb_dir, "master_price_list.xlsx")
    wb_price = openpyxl.Workbook()
    wb_price.remove(wb_price.active)
    for sheet_name in ["Price list", "discount"]:
        sh_dst = wb_price.create_sheet(title=sheet_name)
        if sheet_name in excel_data:
            for row_vals in excel_data[sheet_name]:
                sh_dst.append(row_vals)
        elif sheet_name in wb.sheetnames:
            sh_src = wb[sheet_name]
            for r in range(1, sh_src.max_row + 1):
                row_vals = [sh_src.cell(r, c).value for c in range(1, sh_src.max_column + 1)]
                sh_dst.append(row_vals)
    wb_price.save(master_price_path)
    
    # Export master_discount_list.xlsx
    master_discount_path = os.path.join(wb_dir, "master_discount_list.xlsx")
    wb_disc = openpyxl.Workbook()
    sh_disc_dst = wb_disc.active
    sh_disc_dst.title = "discount"
    if "discount" in excel_data:
        for row_vals in excel_data["discount"]:
            sh_disc_dst.append(row_vals)
    elif "discount" in wb.sheetnames:
        sh_disc_src = wb["discount"]
        for r in range(1, sh_disc_src.max_row + 1):
            row_vals = [sh_disc_src.cell(r, c).value for c in range(1, sh_disc_src.max_column + 1)]
            sh_disc_dst.append(row_vals)
    wb_disc.save(master_discount_path)
    
    # Increment minor version in version.txt AND process_excel_order.py
    ver_path = os.path.join(wb_dir, "version.txt")
    next_ver = CURRENT_VERSION
    try:
        ver_parts = CURRENT_VERSION.split(".")
        ver_parts[-1] = str(int(ver_parts[-1]) + 1)
        next_ver = ".".join(ver_parts)
    except Exception:
        pass
        
    with open(ver_path, "w", encoding="utf-8") as vf:
        vf.write(f"VERSION={next_ver}\n")
        
    # Also update CURRENT_VERSION inside process_excel_order.py
    py_script = os.path.join(wb_dir, "process_excel_order.py")
    try:
        with open(py_script, "r", encoding="utf-8") as pf:
            py_code = pf.read()
        py_code_new = py_code.replace(f'CURRENT_VERSION = "{CURRENT_VERSION}"', f'CURRENT_VERSION = "{next_ver}"')
        with open(py_script, "w", encoding="utf-8") as pf:
            pf.write(py_code_new)
    except Exception as py_ver_err:
        print(f"Note updating py script version: {py_ver_err}")
        
    print(f"Master files exported and Version bumped to v{next_ver}.")
    
    # Call github_api_push.py
    import subprocess
    push_script = os.path.join(wb_dir, "github_api_push.py")
    if not os.path.exists(push_script):
        try:
            import urllib.request
            gpush_url = GITHUB_RAW_BASE + "github_api_push.py"
            urllib.request.urlretrieve(gpush_url, push_script)
            print(f"Auto-downloaded missing {push_script} from GitHub!")
        except Exception as dl_err:
            print(f"Failed to download github_api_push.py: {dl_err}")
            
    res = subprocess.run([sys.executable, push_script], capture_output=True, text=True)
    print("Push Output:\n", res.stdout)
    if res.stderr:
        print("Push Errors:\n", res.stderr)
        
    notice_path = os.path.join(wb_dir, "update_notice.txt")
    if res.returncode == 0 and "[FAILED]" not in res.stdout:
        msg = f"🎉 MASTER PRICE LIST & DISCOUNTS PUSHED TO GITHUB!\n\nNew Version v{next_ver} is now LIVE on GitHub.\nAll team members will automatically receive these updated rates when they import orders or click Update App."
        with open(notice_path, "w", encoding="utf-8") as nf:
            nf.write("STATUS=UPDATED\n")
            nf.write(f"VERSION={next_ver}\n")
            nf.write(f"MSG={msg}")
    else:
        err_detail = res.stderr.strip() if res.stderr else res.stdout.strip()
        msg = f"Error pushing to GitHub:\n{err_detail}"
        with open(notice_path, "w", encoding="utf-8") as nf:
            nf.write("STATUS=ERROR\n")
            nf.write(f"MSG={msg}")

if __name__ == "__main__":
    import argparse
    import traceback
    
    args_workbook = None
    try:
        parser = argparse.ArgumentParser(description="Process Excel Order Sheets")
        parser.add_argument("--import", action="store_true", dest="import_mode", help="Run in order import mode")
        parser.add_argument("--fill", action="store_true", dest="fill_mode", help="Run in template fill mode")
        parser.add_argument("--check-update", action="store_true", dest="check_update_mode", help="Check for updates from GitHub")
        parser.add_argument("--push-masters", action="store_true", dest="push_masters_mode", help="Push updated Price list and discounts to GitHub")
        parser.add_argument("--order", type=str, help="Path to selected order file")
        parser.add_argument("--workbook", type=str, required=True, help="Path to active workbook")
        
        args = parser.parse_args()
        args_workbook = args.workbook
        
        if args.check_update_mode:
            check_for_updates(args.workbook, force_download=True)
        elif args.push_masters_mode:
            push_team_masters(args.workbook)
        elif args.import_mode:
            check_for_updates(args_workbook)
            if not args.order:
                print("Error: --order is required in import mode.")
                sys.exit(1)
            run_import(args.order, args.workbook)
        elif args.fill_mode:
            check_for_updates(args_workbook)
            run_fill(args.workbook)
        else:
            print("Error: Specify --import, --fill, --check-update or --push-masters")
            sys.exit(1)
            
    except Exception as e:
        wb_dir = os.path.dirname(os.path.abspath(args_workbook)) if args_workbook else os.path.dirname(os.path.abspath(__file__))
        log_path = os.path.join(wb_dir, "error_log.txt")
        try:
            with open(log_path, "w", encoding="utf-8") as f:
                f.write("=== Python Process Error Log ===\n")
                f.write(traceback.format_exc())
            print(f"Exception occurred. Error written to: {log_path}")
        except Exception as write_err:
            print(f"Failed to write log file: {write_err}")
        sys.exit(1)
